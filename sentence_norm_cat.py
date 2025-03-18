from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from string_normalizer import TextProcessor
import os
import re

text_processor = TextProcessor()


def normalize_text(text):
    """Normalize text using TextProcessor."""
    return text_processor.process_text(text)


def split_sentences(text):
    """
    Splits text into sentences while preserving full stops, initials, abbreviations, and decimal numbers.
    """
    sentence_pattern = (
        r"(?<!\b[A-Z])(?<!\b[A-Z]\.)(?<!\b[A-Z]\.[A-Z])(?<!\b\d)([.?!])\s+"
    )
    sentences = re.split(sentence_pattern, text)

    result = []
    for i in range(0, len(sentences) - 1, 2):
        result.append(sentences[i] + sentences[i + 1])

    if len(sentences) % 2 == 1:
        result.append(sentences[-1])

    return [sentence.strip() for sentence in result if sentence.strip()]


def categorize_sentence(sentence):
    """Categorize sentences based on word count."""
    word_count = len(sentence.split())
    if word_count <= 4:
        return "Very Short (1-4 words)"
    elif word_count <= 8:
        return "Short (5-8 words)"
    elif word_count <= 11:
        return "Medium (9-11 words)"
    elif word_count <= 15:
        return "Long (12-15 words)"
    else:
        return "Very Long (16+ words)"


# Load data from english_news_articles.xlsx
input_file = "english_news_articles.xlsx"
wb_input = load_workbook(input_file)
sheet = wb_input.active  # Read the first sheet

headers = [str(cell.value).strip().lower() for cell in sheet[1]]

if "domain" not in headers or "text" not in headers:
    raise ValueError("Missing 'Domain' or 'text' columns in the Excel file.")

category_idx = headers.index("domain")
text_idx = headers.index("text")

data = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[category_idx] is None or row[text_idx] is None:
        continue  # Skip empty rows

    category = str(row[category_idx]).strip()
    normalized_text = normalize_text(str(row[text_idx]).strip())
    sentences = split_sentences(normalized_text)

    if category not in data:
        data[category] = {
            "Very Short (1-4 words)": [],
            "Short (5-8 words)": [],
            "Medium (9-11 words)": [],
            "Long (12-15 words)": [],
            "Very Long (16+ words)": [],
        }

    for sentence in sentences:
        sentence_category = categorize_sentence(sentence)
        data[category][sentence_category].append(sentence)

# Load or create output Excel file
output_file = "Categorized_Sentences.xlsx"
if os.path.exists(output_file):
    wb_output = load_workbook(output_file)
else:
    wb_output = Workbook()
    wb_output.remove(wb_output.active)

bold_font = Font(bold=True)
header_row = [
    "Very Short (1-4 words)",
    "Short (5-8 words)",
    "Medium (9-11 words)",
    "Long (12-15 words)",
    "Very Long (16+ words)",
]

for category, sentences_dict in data.items():
    if category not in wb_output.sheetnames:
        ws = wb_output.create_sheet(title=category)
    else:
        ws = wb_output[category]

    # Clear existing content and ensure header row is added
    ws.delete_rows(1, ws.max_row)
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = bold_font

    max_length = max(len(lst) for lst in sentences_dict.values())
    for i in range(max_length):
        row = [
            sentences_dict[key][i] if i < len(sentences_dict[key]) else ""
            for key in header_row
        ]
        ws.append(row)

wb_output.save(output_file)
print("Sentences have been split, categorized, and saved successfully.")
